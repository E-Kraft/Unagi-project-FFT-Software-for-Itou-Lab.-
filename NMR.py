import numpy as np #FFTに使用します
import struct #バイナリの読み込みに使用します

#-------------------今後追加予定の項目------------------------------------
#ゼロパディングモードの実装
#窓関数のテスト
#NMR信号のファイルパス一覧を返す関数

#ピーク計算&重心計算モード(積分機能の実装)
#生データをcsvあるいはエクセルファイルに保存する機能
#フーリエ変換データをcsvあるいはエクセルファイルに保存する機能
#fft合成&merge


# 各項目のアドレスとデータ型
block_def = {
    'cos'               : {'offset':    10, 'type': ' 8192f'},
    'sin'               : {'offset': 32788, 'type': '8192f'},
    'inte'              : {'offset': 65556, 'type': 'd'},
    'sigma'             : {'offset': 65564, 'type': 'd'},
    'dt0'               : {'offset': 65572, 'type': 'd'},
    'dt1'               : {'offset': 65580, 'type': 'd'},
    'variable'          : {'offset': 65588, 'type': '27d'},
    'field'             : {'offset': 65804, 'type': 'd'},
    'temp'              : {'offset': 65812, 'type': 'd'},
    'Frec'              : {'offset': 65820, 'type': 'd'},
    'TXLVL'             : {'offset': 65828, 'type': 'f'},
    'RFATT'             : {'offset': 65832, 'type': 'f'},
    'Att3'              : {'offset': 65836, 'type': 'f'},
    'RFSW'              : {'offset': 65840, 'type': 'h'},
    'GAIN'              : {'offset': 65842, 'type': 'd'},
    'Phas'              : {'offset': 65850, 'type': 'd'},
    'LPF1'              : {'offset': 65858, 'type': 'd'},
    'X'                 : {'offset': 65876, 'type': 'd'},
    'Y'                 : {'offset': 65884, 'type': 'd'},
    'Z'                 : {'offset': 65892, 'type': 'd'},
    'isDouble'          : {'offset': 65900, 'type': 'h'},
    'blank'             : {'offset': 65902, 'type': 'd'},
    'cpn'               : {'offset': 65910, 'type': 'h'},
    'cpw'               : {'offset': 65912, 'type': 'd'},
    'cpi'               : {'offset': 65920, 'type': 'd'},
    'fpw'               : {'offset': 65928, 'type': 'd'},
    'spw'               : {'offset': 65936, 'type': 'd'},
    'tJ'                : {'offset': 65944, 'type': 'd'},
    't2'                : {'offset': 65952, 'type': 'd'},
    'ADOFF'             : {'offset': 65960, 'type': 'd'},
    'Reserve1'          : {'offset': 65968, 'type': 'd'},
    'OSCILLO'           : {'offset': 65976, 'type': 'h'},
    'PowLevel'          : {'offset': 65978, 'type': 'h'},
    'QPSKComb'          : {'offset': 65980, 'type': 'h'},
    'QPSK1st'           : {'offset': 65982, 'type': 'h'},
    'QPSK2nd'           : {'offset': 65984, 'type': 'h'},
    'ADTrigLocation'    : {'offset': 65986, 'type': 'h'},
    'EXTTRIG'           : {'offset': 65988, 'type': 'h'},
    'BlankIsLoopTime'   : {'offset': 65990, 'type': 'h'},
    'UseComb'           : {'offset': 65992, 'type': 'h'},
    'QPSKRX'            : {'offset': 65994, 'type': 'h'},
    'mDate'             : {'offset': 65996, 'type': '20c'},
    'mTime'             : {'offset': 66026, 'type': '30c'},
    'real'              : {'offset': 66056, 'type': 'd'},
    'mTitle'            : {'offset': 66064, 'type': '32c'},
    'wavesize'          : {'offset': 66096, 'type': 'i'}
}

#測定データのヘッダー(10バイト,2バイトずつ)
check_header = np.array([1,8192,0,0,0])

#生データクラスを定義
class raw:
    def __init__(self,path):
        #ファイルのパス
        self.path = path

        #メインの情報
        self.cos = np.array([]) #元の信号
        self.sin = np.array([]) #位相を90°ズラした信号
        self.sigma = 0 #積算回数
        self.dt0 = 0.0 #サンプリング周期(cos)
        self.dt1 = 0.0 #サンプリング周期(sin)
        self.Frec = 0.0 #参照周波数
        self.TXLVL = 0.0 #OUTPUT LEVEL
        self.GAIN = 0.0 #GAIN
        self.LPF1 = 0.0 #LPFのカットオフ周波数
        self.Phas = 0.0 #phase
        self.wavesize = 0 #サンプリング数

        #modulator
        self.RFATT = 0.0
        self.Att3 = 0.0
        self.RFSW = 0


        #T1T2info
        self.isDouble = 0
        self.blank = 0.0
        self.cpn = 0
        self.cpw = 0.0
        self.cpi = 0.0
        self.fpw = 0.0
        self.spw = 0.0
        self.tJ = 0.0
        self.t2 = 0.0
        self.ADOFF = 0.0
        self.Reserve1 = 0.0
        self.OSCILLO = 0
        self.PowLevel = 0
        self.QPSKComb = 0
        self.QPSK1st = 0
        self.QPSK2nd = 0
        self.ADTrigLocation = 0
        self.EXTTRIG = 0
        self.BlankIsLoopTime = 0
        self.UseComb = 0
        self.QPSKRX = 0

        #others
        self.inte = 0.0
        self.field = 0.0
        self.temp = 0.0
        self.X = 0.0
        self.Y = 0.0
        self.Z = 0.0
        self.real = 0.0

        #variable
        self.variable = []

        #メタデータ
        self.mDate = "" #測定日
        self.mTime = "" #測定時刻
        self.mTitle = "" #タイトル

    #積算回数sigmaで信号データを割り算して正規化
    #cosの正規化
    def normalize_cos(self):
        normalized_cos = self.cos / self.sigma
        return normalized_cos

    #sinの正規化
    def normalize_sin(self):
        normalized_sin = self.sin / self.sigma
        return normalized_sin

    def main(self):
        main_list = {"積算回数":self.sigma,"サンプリング周期(cos)":self.dt0,"サンプリング周期(sin)":self.dt1,"Frequency":self.Frec,"OUTPUT LEVEL":self.TXLVL,
                    "GAIN":self.GAIN,"LPF [Hz]":self.LPF1,"Phase":self.Phas,"サンプリング数":self.wavesize}
        return main_list

    def modulator(self):
        modulator_list = {"RFATT":self.RFATT,"Att3":self.Att3,"RFSW":self.RFSW}
        return modulator_list

    def T1T2info(self):
        T1T2info_list = {"isDouble":self.isDouble,"blank":self.blank,"cpn":self.cpn,"cpw":self.cpw,"cpi":self.cpi,"fpw":self.fpw,"spw":self.spw,"tJ":self.tJ,
                        "t2":self.t2,"ADOFF":self.ADOFF,"Reserve1":self.Reserve1,"OSCILLO":self.OSCILLO,"PowLevel":self.PowLevel,"QPSKComb":self.QPSKComb,
                        "QPSK1st":self.QPSK1st,"QPSK2nd":self.QPSK2nd,"ADTrigLocation":self.ADTrigLocation,"EXTTRIG":self.EXTTRIG,"BlankIsLoopTime":self.BlankIsLoopTime,
                        "UseComb":self.UseComb,"QPSKRX":self.QPSKRX}
        return T1T2info_list

    def others(self):
        others_list = {"inte":self.inte,"field":self.field,"temp":self.temp,"X":self.X,"Y":self.Y,"Z":self.Z,"real":self.real}
        return others_list

    #これだけリストで返すので注意
    def var(self):
        variable_list = self.variable
        return variable_list

    def meta(self):
        meta_list = {"測定日":self.mDate,"測定時刻":self.mTime,"タイトル":self.mTitle}
        return meta_list

    #時間列の作成
    def time_column(self):
        time = np.arange(self.wavesize)*self.dt0
        return time

    #FFTの実行
    def FFT(self,zeropoint):
        #ゼロパディングを実行
        signal_complex = self.normalize_sin() + 1j*self.normalize_cos()
        signal_complex = np.concatenate([signal_complex[zeropoint:], np.zeros(zeropoint,dtype=complex)])

        #FFTを実行
        fft_result = np.fft.fft(signal_complex)

        #順番をマイナスからプラスに
        fft_result = np.fft.fftshift(fft_result)
        fft_re = fft_result.real
        fft_im = fft_result.imag
        return (fft_re,fft_im,fft_result)

    #周波数軸の生成
    def calc_freq(self):
        #単位をkHzに
        freq_column = np.fft.fftfreq(self.wavesize,self.dt0)/1000
        freq_column = np.fft.fftshift(freq_column)
        return freq_column

def header_check(path):
    with open(path,'rb') as f:
        data = f.read()
    if len(data) < 10: #10バイト未満の場合は排除
        return False

    header = np.array(struct.unpack_from('5h',data,0))
    if all(header == check_header):
        return True
    else:
        return False

#生データをrawクラスのインスタンスとしてインポート
def import_rawdata(path):
    try:
        with open(path,'rb') as f:
            raw_data = raw(path)
            data = f.read()
            #ヘッダーの読み取り
            if len(data) < 10: #10バイト未満の場合は排除
                raise Exception('Not NMR signal file!') #NMRの信号ファイルではなかった場合例外を発生

            header = np.array(struct.unpack_from('5h',data,0))
            if all(header == check_header):
                for name ,dict in block_def.items():
                    cache = np.array(struct.unpack_from(dict['type'],data,dict['offset'])) #値はタプルで返される
                    #タプルの要素が一つなら数値に戻す
                    if len(cache) == 1:
                        cache = cache[0]
                    setattr(raw_data,name,cache) #nameのプロパティに代入
                #バイト列から文字列に修正
                raw_data.mDate = bytes(raw_data.mDate).decode().strip()
                raw_data.mTime = bytes(raw_data.mTime).decode().strip()
                raw_data.mTitle = bytes(raw_data.mTitle).decode().strip()
                return  raw_data
            else:
                raise Exception('Not NMR signal file!') #NMRの信号ファイルではなかった場合例外を発生
    except Exception as e:
        print(f"エラーが発生しました : {e}") #この関数内で起こった例外はすべてここで処理される

#位相調整(nは整数を入れる)
def phase_shift(fft_result,n,base_angle = np.pi / 100):
    shifted_data = fft_result*np.exp(1j*n*base_angle)
        #実部と虚部の取り出し
    shifted_re = shifted_data.real
    shifted_im = shifted_data.imag
    return (shifted_re,shifted_im,shifted_data)

#自動位相調整(base_angleを単位として位相変化)
def auto_const(fft_result,base_angle = np.pi / 100):
    #位相の変化一覧を作っておく
    phase_list = np.arange(-np.pi/2 ,np.pi/2,base_angle)
    #積分値を入れるリストを作っておく
    inte = np.zeros(len(phase_list))

    for index,phase in enumerate(phase_list):
        #位相回転
        shifted_data = fft_result*np.exp(1j*phase)
        #積分値
        inte[index] = sum(shifted_data.real)

    #積分値最大のインデックスを検索
    auto_phase = phase_list[np.argmax(inte)]
    #積分値最大の位相でシフト
    shifted_data = fft_result*np.exp(1j*auto_phase)
    #実部と虚部の取り出し
    shifted_re = shifted_data.real
    shifted_im = shifted_data.imag
    return (shifted_re,shifted_im,shifted_data)



if __name__ == "__main__":
    # ここは直接実行されたときだけ実行
    # モジュールとしてインポートされた場合は実行されない(テストに用いる)
    # __name__にはメインスクリプトとして実行された場合は__main__、モジュールとしてインポートされた場合はファイル名が代入される

    def save_excel(file_path):
        import sys #システム諸々で使用します
        import openpyxl  as opx #エクセルの書き込みに使用
        from openpyxl.styles import Font, PatternFill

        #変数の定義
        zeropoint = 1400 #ゼロポイント
        n_ind= 3441 #切り取り範囲(負)
        p_ind = 4752 #切り取り範囲(正)
        base_angle = np.pi / 100 #基本の位相回転角(フーリエ変換ver2.5参照)　optionで選択できるようにするかも

        #-------------実行部分---------------------

        #データの読み込み
        raw_data = import_rawdata(file_path)

        #周波数軸を生成
        freq_column= raw_data.calc_freq()
        #FFT実行
        (fft_re,fft_im,fft_result) = raw_data.FFT(zeropoint)

        #位相自動調整
        (shifted_re,shifted_im,_)= auto_const(fft_result)


        #-------------エクセルに書き込み---------------------

        wb = opx.Workbook() # workbookの作成
        ws = wb.worksheets[0]

        cell = ws.cell(row=1, column=1, value="Frequency [kHz]")
        cell.fill = PatternFill("solid", fgColor="1F497D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,v in enumerate(freq_column[n_ind:p_ind],start=1):
            ws.cell(row=i+1, column=1, value=v)

        cell = ws.cell(row=1, column=2, value="Real")
        cell.fill = PatternFill("solid", fgColor="1F497D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,v in enumerate(fft_re[n_ind:p_ind],start=1):
            ws.cell(row=i+1, column=2, value=v)

        cell = ws.cell(row=1, column=3, value="Image")
        cell.fill = PatternFill("solid", fgColor="1F497D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,v in enumerate(fft_im[n_ind:p_ind],start=1):
            ws.cell(row=i+1, column=3, value=v)

        cell = ws.cell(row=1, column=4, value="shifted_Real")
        cell.fill = PatternFill("solid", fgColor="1F497D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,v in enumerate(shifted_re[n_ind:p_ind],start=1):
            ws.cell(row=i+1, column=4, value=v)

        cell = ws.cell(row=1, column=5, value="shifted_Image")
        cell.fill = PatternFill("solid", fgColor="1F497D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,v in enumerate(shifted_im[n_ind:p_ind],start=1):
            ws.cell(row=i+1, column=5, value=v)

        cell = ws.cell(row=1, column=6, value="main")
        cell.fill = PatternFill("solid", fgColor="C0504D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,(key,value) in enumerate(raw_data.main().items(),start=1):
            ws.cell(row=2*i, column=6, value=key)
            ws.cell(row=2*i+1, column=6, value=value)

        cell = ws.cell(row=1, column=7, value="modulator")
        cell.fill = PatternFill("solid", fgColor="C0504D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,(key,value) in enumerate(raw_data.modulator().items(),start=1):
            ws.cell(row=2*i, column=7, value=key)
            ws.cell(row=2*i+1, column=7, value=value)

        cell = ws.cell(row=1, column=8, value="T1T1info")
        cell.fill = PatternFill("solid", fgColor="C0504D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,(key,value) in enumerate(raw_data.T1T2info().items(),start=1):
            ws.cell(row=2*i, column=8, value=key)
            ws.cell(row=2*i+1, column=8, value=value)

        cell = ws.cell(row=1, column=9, value="others")
        cell.fill = PatternFill("solid", fgColor="C0504D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,(key,value) in enumerate(raw_data.others().items(),start=1):
            ws.cell(row=2*i, column=9, value=key)
            ws.cell(row=2*i+1, column=9, value=value)

        cell = ws.cell(row=1, column=10, value="variable")
        cell.fill = PatternFill("solid", fgColor="C0504D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,v in enumerate(raw_data.var(),start=1):
            ws.cell(row=i+1, column=10, value=v)

        cell = ws.cell(row=1, column=11, value="meta_data")
        cell.fill = PatternFill("solid", fgColor="C0504D")
        cell.font = Font(color="FFFFFF", bold=True)
        for i,(key,value) in enumerate(raw_data.meta().items(),start=1):
            ws.cell(row=2*i, column=11, value=key)
            ws.cell(row=2*i+1, column=11, value=value)

        #列幅調整
        column_width = {"A":16, "B":14, "C":14, "D":14, "E":14, "F":19.5, "G":11, "H":15, "I":13, "J":9, "K":35}

        for column , width in column_width.items():
            ws.column_dimensions[column].width = width

        ws.title = (file_path) # シート名の変更
        wb.save("fft_shift.xlsx") # Excelファイルの保存

    file_path = '10k0613T2F.1010'
    save_excel(file_path)
