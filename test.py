import NMR #生データ読み込み用のクラスと関数
import matplotlib.pyplot as plt #グラフ
import numpy as np #フーリエ変換など
import openpyxl  as opx #エクセルの書き込みに使用

#変数の定義
zeropoint = 1400 #ゼロポイント
n_ind= 3441 #切り取り範囲(負)
p_ind = 4752 #切り取り範囲(正)
base_angle = np.pi / 100 #基本の位相回転角(フーリエ変換ver2.5参照)　optionで選択できるようにするかも

#-------------実行部分---------------------

#データの読み込み
raw_data = NMR.import_rawdata('10k0613T2F.1010')

#周波数軸を生成
freq_column= raw_data.calc_freq()
#FFT実行
(fft_re,fft_im,fft_result) = raw_data.FFT(zeropoint)

#位相自動調整
(shifted_re,shifted_im,shifted_fft)= NMR.auto_const(fft_result)


#-------------エクセルに書き込み---------------------

wb = opx.Workbook() # workbookの作成
ws = wb.worksheets[0]

ws.cell(row=1, column=1, value="Frequency [kHz]")
for i,v in enumerate(freq_column[n_ind:p_ind],start=1):
    ws.cell(row=i+1, column=1, value=v)

ws.cell(row=1, column=2, value="Real")
for i,v in enumerate(fft_re[n_ind:p_ind],start=1):
    ws.cell(row=i+1, column=2, value=v)

ws.cell(row=1, column=3, value="Image")
for i,v in enumerate(fft_im[n_ind:p_ind],start=1):
    ws.cell(row=i+1, column=3, value=v)

ws.cell(row=1, column=4, value="shifted_Real")
for i,v in enumerate(shifted_re[n_ind:p_ind],start=1):
    ws.cell(row=i+1, column=4, value=v)

ws.cell(row=1, column=5, value="shifted_Image")
for i,v in enumerate(shifted_im[n_ind:p_ind],start=1):
    ws.cell(row=i+1, column=5, value=v)


wb.save("fft_shift.xlsx") # Excelファイルの保存
